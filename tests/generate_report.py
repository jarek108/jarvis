import pandas as pd
import json
import os
import argparse
import pickle
import time
import traceback
import webbrowser
import re
import wave
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

from test_utils.reporting import GDriveAssetManager

SCOPES = ['https://www.googleapis.com/auth/drive.file']

def load_json(path):
    if not os.path.exists(path): return []
    with open(path, "r", encoding="utf-8") as f:
        try: return json.load(f)
        except: return []

def get_gdrive_service():
    try:
        creds = None
        tests_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(tests_dir)
        pickle_path = os.path.join(project_root, 'token.pickle')
        creds_path = os.path.join(project_root, 'credentials.json')
        if os.path.exists(pickle_path):
            with open(pickle_path, 'rb') as token: creds = pickle.load(token)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try: creds.refresh(Request())
                except:
                    flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
                    creds = flow.run_local_server(port=0)
            else:
                flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
                creds = flow.run_local_server(port=0)
            with open(pickle_path, 'wb') as token: pickle.dump(creds, token)
        return build('drive', 'v3', credentials=creds)
    except Exception as e:
        print(f"❌ GDrive Auth Error: {e}"); return None

def infer_detailed_name(model_id):
    if not model_id or model_id == "N/A": return model_id
    if "_" in model_id: return " + ".join([infer_detailed_name(p) for p in model_id.split("_")])
    name = model_id.upper()
    if "#CTX=" not in name:
        if "OL_" in name: name += "#CTX=4096"
        elif "VL_" in name: name += "#CTX=16384"
    return name

def generate_excel(upload=True, upload_outputs=False, session_dir=None):
    try:
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if session_dir:
            artifacts_dir = os.path.abspath(session_dir)
            session_id = os.path.basename(artifacts_dir)
        else:
            artifacts_dir = os.path.join(project_root, "tests", "artifacts")
            session_id = "LATEST"
            
        ts_part = session_id.replace("RUN_", "") if session_id.startswith("RUN_") else time.strftime("%Y-%m-%d_%H-%M-%S")
        output_path = os.path.join(artifacts_dir, f"Jarvis_Benchmark_Report_{ts_part}.xlsx")
        service = get_gdrive_service() if upload else None
        asset_mgr = GDriveAssetManager(service) if service else None
        
        master_id = asset_mgr.get_folder_id("Jarvis") if asset_mgr else None
        inputs_id = asset_mgr.get_folder_id("Inputs", master_id) if asset_mgr else None
        outputs_root_id = asset_mgr.get_folder_id("Outputs", master_id) if asset_mgr else None
        session_out_id = asset_mgr.get_folder_id(session_id, outputs_root_id) if asset_mgr else None
        session_out_link = asset_mgr.get_folder_link(session_id, outputs_root_id) if asset_mgr else None

        all_data = {}
        for f in os.listdir(artifacts_dir):
            if f.endswith(".json") and not f.startswith("system_info"):
                all_data[f.replace(".json", "").replace("latest_", "")] = load_json(os.path.join(artifacts_dir, f))

        def get_link(local_path, folder_id=None):
            if not local_path: return "N/A"
            abs_p = os.path.abspath(os.path.join(project_root, local_path))
            fname = os.path.basename(abs_p)
            if asset_mgr and folder_id:
                cache = asset_mgr.file_cache.get(folder_id, {})
                if fname in cache:
                    label = "▶️ Play" if ".wav" in fname else ("👁️ View" if "." in fname else "📄 Log")
                    return f'=HYPERLINK("{cache[fname]}", "{label}")'
            return f'=HYPERLINK("{abs_p}", "📁 Local File")'

        def r3(val):
            try: return round(float(val), 3) if val is not None else 0
            except: return 0

        sheets = {}
        
        # 1. Summary Sheet
        sys_info_path = os.path.join(artifacts_dir, "system_info.yaml")
        if os.path.exists(sys_info_path):
            import yaml
            with open(sys_info_path, "r", encoding="utf-8") as f: sys_info = yaml.safe_load(f)
            run_val = f'=HYPERLINK("{session_out_link}", "{session_id}")' if session_out_link else session_id
            summary_rows = [
                {"Category": "Session", "Metric": "Run ID", "Value": run_val},
                {"Category": "Session", "Metric": "Original Timestamp", "Value": sys_info.get("timestamp", "N/A")},
                {"Category": "Session", "Metric": "Report Generated", "Value": time.strftime("%Y-%m-%d %H:%M:%S")}
            ]
            for k, v in sys_info.get("host", {}).items():
                if isinstance(v, dict):
                    for sk, sv in v.items(): summary_rows.append({"Category": "Host", "Metric": f"{k.title()} {sk.title()}", "Value": str(sv)})
                else: summary_rows.append({"Category": "Host", "Metric": k.replace("_", " ").title(), "Value": str(v)})
            sheets["Summary"] = pd.DataFrame(summary_rows)

        # 2. Dynamic Domain Sheets
        for domain_raw, data in all_data.items():
            domain = domain_raw.upper()
            if not data: continue
            print(f"📊 Processing {domain} sheet...")
            
            node_columns = []
            for entry in data:
                for s in entry.get('scenarios', []):
                    for nid, metrics in s.get('node_metrics', {}).items():
                        for m_key in metrics.keys():
                            col_name = f"{nid} ({m_key.upper()})"
                            if col_name not in node_columns: node_columns.append(col_name)
            node_columns.sort()

            rows = []
            for entry in data:
                loadout_name = entry.get('loadout', 'N/A')
                # If there are NO successful scenarios, we still want to show the Lifecycle failure
                scenarios = entry.get('scenarios', [])
                valid_scenarios = [s for s in scenarios if s.get('name') not in ["SETUP", "LIFECYCLE"]]
                
                if not valid_scenarios:
                    # Look for the failure message in the LIFECYCLE entry
                    lc = next((s for s in scenarios if s.get('name') == "LIFECYCLE"), None)
                    if lc:
                        rows.append({
                            "Loadout": infer_detailed_name(lc.get('detailed_model') or loadout_name),
                            "Scenario": "CRITICAL_FAILURE",
                            "Status": "FAILED",
                            "Exec": 0,
                            "Setup": r3(lc.get('setup_time', 0)),
                            "Cleanup": r3(lc.get('cleanup_time', 0)),
                            "V_BG": 0, "V_Static": 0, "V_Peak": 0,
                            "Response": f"LIFECYCLE ERROR: {lc.get('result', 'Unknown error')}"
                        })

                for s in valid_scenarios:
                    row = {
                        "Loadout": infer_detailed_name(s.get('detailed_model') or loadout_name),
                        "Scenario": s.get('name'),
                        "Status": s.get('status'),
                        "Exec": r3(s.get('duration', 0)),
                        "Setup": r3(s.get('setup_time', 0)),
                        "Cleanup": r3(s.get('cleanup_time', 0)),
                        "V_BG": r3(s.get('vram_background', 0)),
                        "V_Static": r3(s.get('vram_static', 0)),
                        "V_Peak": r3(s.get('vram_peak', 0))
                    }
                    
                    for col in node_columns:
                        nid, m_key = col.split(" ("); m_key = m_key[:-1].lower()
                        val = s.get('node_metrics', {}).get(nid, {}).get(m_key, "")
                        row[col] = r3(val) if isinstance(val, (int, float)) else val

                    # Fuzzy match for log link
                    log_link = "N/A"
                    role_map = {"stt": "stt", "tts": "tts", "llm": "llm", "sts_voice": "llm", "hybrid_chat": "llm", "vlm": "llm"}
                    target_role = role_map.get(domain_raw, "llm")
                    log_files = [f for f in os.listdir(artifacts_dir) if f.startswith(f"svc_{target_role}") and f.endswith(".log")]
                    match_found = next((lf for lf in log_files if entry.get('loadout_id', '').lower() in lf.lower()), None)
                    if not match_found and log_files: match_found = log_files[0]
                    if match_found: log_link = get_link(os.path.join(artifacts_dir, match_found), session_out_id)

                    row.update({
                        "Input": get_link(s.get('input_file'), inputs_id),
                        "Output": get_link(s.get('output_file'), session_out_id),
                        "Logs": log_link,
                        "Prompt": str(s.get('input_text', "N/A"))[:500],
                        "Response": str(s.get('llm_text', "N/A"))[:1000]
                    })
                    rows.append(row)
            if rows: sheets[domain] = pd.DataFrame(rows)

        # 3. Write Excel
        from openpyxl.styles import Font, PatternFill
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for name, df in sheets.items():
                df.to_excel(writer, sheet_name=name, index=False)
                worksheet = writer.sheets[name]; last_row = len(df) + 1
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for idx, col in enumerate(df.columns):
                    cell = worksheet.cell(row=1, column=idx+1)
                    cell.fill = header_fill; cell.font = header_font
                    col_letter = get_column_letter(idx + 1)
                    if col == "Loadout": width = 40
                    elif any(x in col.lower() for x in ["input", "output", "logs"]): width = 12
                    elif any(x in col.lower() for x in ["prompt", "response"]): width = 50
                    else: width = 15
                    worksheet.column_dimensions[col_letter].width = width
                    if col in ["Status"]:
                        worksheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{last_row}", FormulaRule(formula=[f'{col_letter}2="PASSED"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
                        worksheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{last_row}", FormulaRule(formula=[f'{col_letter}2="FAILED"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
        print(f"📊 Excel Report Generated: {output_path}"); return output_path
    except Exception as e:
        print(f"❌ Excel Error: {e}"); traceback.print_exc(); return None

def generate_and_upload_report(session_dir, upload_report=True, upload_outputs=False, open_browser=True):
    path = generate_excel(upload=upload_report, upload_outputs=upload_outputs, session_dir=session_dir)
    if not path: return None
    if upload_report:
        service = get_gdrive_service()
        if service:
            manager = GDriveAssetManager(service)
            master_id = manager.get_folder_id("Jarvis")
            print(f"📤 Uploading report to GDrive master folder...")
            link = manager.sync_file(path, master_id, overwrite=False)
            if link:
                print(f"✅ GDrive Link: {link}\n🌐 Opening report in browser...")
                if open_browser: webbrowser.open(link)
                return link
    return path

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--upload-report", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--dir", type=str)
    args = parser.parse_args()
    generate_and_upload_report(session_dir=args.dir, upload_report=args.upload_report)

if __name__ == "__main__":
    main()
